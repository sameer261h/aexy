"""Questionnaire import and analysis service.

Handles Excel questionnaire parsing, question extraction,
and ReminderSuggestion generation for compliance tracking.

Column Detection Strategy (3-tier):
1. Exact alias match — fast, no LLM cost, handles ~90% of formats
2. Fuzzy substring match — catches partial/compound headers
3. LLM fallback — when auto-detection fails, asks the LLM to map columns
"""

import io
import json
import logging
import re
from collections import Counter
from dataclasses import dataclass
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from aexy.models.questionnaire import QuestionnaireQuestion, QuestionnaireResponse
from aexy.models.reminder import ReminderSuggestion

logger = logging.getLogger(__name__)

# Keywords used to detect frequency-type responses
FREQUENCY_KEYWORDS = {"monthly", "quarterly", "half yearly", "annually", "yearly", "weekly", "daily", "never"}

# Keywords that indicate a review/audit/training domain for yearly suggestions
YEARLY_INDICATOR_KEYWORDS = {
    "review", "audit", "training", "policy", "certification",
    "assessment", "testing", "drill", "inspection", "evaluation",
}

# Mapping from frequency answer text to reminder frequency values
FREQUENCY_MAP = {
    "daily": "daily",
    "weekly": "weekly",
    "monthly": "monthly",
    "quarterly": "quarterly",
    "half yearly": "semi_annual",
    "semi-annual": "semi_annual",
    "semi-annually": "semi_annual",
    "biannual": "semi_annual",
    "annually": "yearly",
    "yearly": "yearly",
}

# Headers to look for when detecting the questionnaire sheet (expanded)
HEADER_PATTERNS = {
    "domain", "question", "response", "control", "requirement",
    "answer", "category", "section", "assessment", "finding",
}

# Column name variations for auto-detection (expanded)
COLUMN_ALIASES = {
    "serial_number": {
        "sr.no", "sr no", "s.no", "s no", "serial", "sl.no", "sl no",
        "#", "no", "no.", "id", "ref", "ref.", "reference", "item",
        "item no", "item no.", "row", "row no", "row no.",
    },
    "domain": {
        "domain", "category", "control domain", "area", "section",
        "topic", "subject", "pillar", "theme", "group", "module",
        "control area", "control category", "control family",
        "security domain", "assessment area", "focus area",
        "framework domain", "capability", "process area",
    },
    "question": {
        "question", "questions", "query", "requirement", "control",
        "control objective", "control description", "assessment criteria",
        "criteria", "check", "checkpoint", "test", "test case",
        "finding", "observation", "item description", "description",
        "control statement", "control requirement", "audit question",
        "compliance requirement", "security requirement",
        "assessment question", "evaluation criteria", "measure",
        "indicator", "parameter", "standard", "clause",
    },
    "response": {
        "response", "answer", "reply", "status", "your response",
        "your answer", "vendor response", "client response",
        "organisation response", "organization response",
        "compliance status", "maturity level", "maturity",
        "assessment result", "result", "rating", "score",
        "current state", "current status", "implementation status",
        "finding", "vendor answer", "auditee response",
        "self assessment", "self-assessment",
    },
    "possible_responses": {
        "possible responses", "options", "possible answers", "choices",
        "expected response", "response options", "allowed values",
        "valid responses", "answer options", "scale", "rating scale",
        "maturity scale", "response choices", "acceptable answers",
        "expected values", "possible values",
    },
    "explanation": {
        "explanation", "comments", "remarks", "notes", "details",
        "justification", "evidence", "evidence reference",
        "supporting evidence", "additional comments", "additional info",
        "additional information", "rationale", "reason", "observation",
        "findings", "gap", "gap description", "recommendation",
        "remediation", "action required", "corrective action",
        "mitigation", "artifacts", "reference", "references",
    },
}

# Fuzzy keywords for substring-based matching (tier 2)
FUZZY_KEYWORDS = {
    "serial_number": ["serial", "sr.", "sl.", "item no", "ref"],
    "domain": ["domain", "category", "section", "area", "control fam"],
    "question": ["question", "requirement", "control", "criteria", "check"],
    "response": ["response", "answer", "status", "result", "rating", "maturity"],
    "possible_responses": ["option", "choice", "possible", "scale", "allowed"],
    "explanation": ["comment", "remark", "note", "evidence", "justification", "explanation"],
}

# LLM prompt for column mapping fallback
LLM_COLUMN_MAPPING_PROMPT = """You are analyzing a spreadsheet header row from a compliance/security questionnaire.

The header row contains these column values (0-indexed):
{headers}

Map each column to one of these field types (or null if it doesn't match any):
- "serial_number" — Serial/row number, ID, reference number
- "domain" — Category, domain, section, topic, control family
- "question" — The actual question, requirement, control statement, criteria
- "response" — The answer, compliance status, maturity level, result
- "possible_responses" — Allowed answer options, rating scale
- "explanation" — Comments, evidence, justification, notes, remarks

Return ONLY valid JSON with this exact format (no markdown, no explanation):
{{"serial_number": <col_index_or_null>, "domain": <col_index_or_null>, "question": <col_index_or_null>, "response": <col_index_or_null>, "possible_responses": <col_index_or_null>, "explanation": <col_index_or_null>}}

Rules:
- "question" is REQUIRED — every questionnaire has a question/requirement column
- Each column index can only be used once
- Use null for fields that have no matching column
- Prefer the most specific match if ambiguous"""


@dataclass
class DedupEntry:
    """Info about the existing item a question duplicates."""
    item_id: str
    item_type: str  # "suggestion" | "reminder" | "question"
    title: str
    domain: str | None = None


@dataclass
class SkippedQuestion:
    """A question that was skipped during suggestion generation."""
    question_text: str
    domain: str | None
    reason: str  # "header" | "blank" | "negative" | "duplicate" | "no_suggestion"
    duplicate_of_id: str | None = None
    duplicate_of_type: str | None = None
    duplicate_of_title: str | None = None


class QuestionnaireServiceError(Exception):
    """Base error for questionnaire service."""
    pass


class QuestionnaireService:
    """Service for importing and analyzing questionnaires."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self._llm_gateway = None

    def _get_llm_gateway(self):
        """Lazy-load the LLM gateway (may be None if not configured)."""
        if self._llm_gateway is None:
            try:
                from aexy.llm.gateway import get_llm_gateway
                self._llm_gateway = get_llm_gateway()
            except Exception:
                logger.debug("LLM gateway not available for column mapping fallback")
        return self._llm_gateway

    # =========================================================================
    # Import
    # =========================================================================

    async def import_questionnaire(
        self,
        workspace_id: str,
        filename: str,
        file_bytes: bytes,
        uploaded_by_id: str,
    ) -> tuple[QuestionnaireResponse, list[QuestionnaireQuestion]]:
        """Import a questionnaire from Excel bytes.

        Returns the created QuestionnaireResponse and list of QuestionnaireQuestion records.
        """
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            raise QuestionnaireServiceError(f"Failed to read Excel file: {e}")

        # Extract metadata from Document Summary sheet (first sheet usually)
        metadata = self._extract_metadata(wb)
        partner_name = metadata.get("partner_name") or metadata.get("company") or metadata.get("organization")
        assessment_year = metadata.get("year") or metadata.get("assessment_year")
        title = metadata.get("title") or f"Questionnaire Import - {filename}"

        # Find the questions sheet
        questions_sheet = self._find_questions_sheet(wb)
        if questions_sheet is None:
            raise QuestionnaireServiceError(
                "Could not find a sheet with questionnaire data. "
                "Expected columns: Domain, Question, Response"
            )

        # Detect column mapping from header row (3-tier: exact → fuzzy → LLM)
        col_map = await self._detect_columns_with_fallback(questions_sheet, workspace_id)
        if "question" not in col_map:
            raise QuestionnaireServiceError(
                "Could not detect 'Question' column in the sheet. "
                "Please ensure the header row contains a 'Question' column."
            )

        # Create the questionnaire response record
        questionnaire = QuestionnaireResponse(
            id=str(uuid4()),
            workspace_id=workspace_id,
            title=title,
            partner_name=partner_name,
            assessment_year=assessment_year,
            source_filename=filename,
            status="uploaded",
            extra_metadata=metadata,
            uploaded_by_id=uploaded_by_id,
        )
        self.db.add(questionnaire)

        # Parse rows into questions
        questions = []
        current_domain = None

        for row_idx, row in enumerate(questions_sheet.iter_rows(min_row=col_map.get("header_row", 1) + 1, values_only=False), start=col_map.get("header_row", 1) + 1):
            row_values = [cell.value for cell in row]

            # Skip completely empty rows
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_values):
                continue

            question_text = self._get_col_value(row_values, col_map, "question")
            if not question_text:
                continue

            domain = self._get_col_value(row_values, col_map, "domain")
            response_text = self._get_col_value(row_values, col_map, "response")
            possible_responses = self._get_col_value(row_values, col_map, "possible_responses")
            explanation = self._get_col_value(row_values, col_map, "explanation")
            serial_number = self._get_col_value(row_values, col_map, "serial_number")

            # Track current domain (domains often span multiple rows)
            if domain:
                current_domain = str(domain).strip()

            # Detect section headers: rows with domain/question but no possible_responses and no response
            is_header = self._is_section_header(question_text, response_text, possible_responses)

            # Classify response type
            response_type = self._classify_response_type(possible_responses, response_text)

            q = QuestionnaireQuestion(
                id=str(uuid4()),
                questionnaire_response_id=questionnaire.id,
                serial_number=str(serial_number).strip() if serial_number else None,
                domain=current_domain,
                question_text=str(question_text).strip(),
                response_text=str(response_text).strip() if response_text else None,
                possible_responses=str(possible_responses).strip() if possible_responses else None,
                explanation=str(explanation).strip() if explanation else None,
                is_section_header=is_header,
                response_type=response_type,
                source_row=row_idx,
            )
            questions.append(q)
            self.db.add(q)

        questionnaire.total_questions = len(questions)
        await self.db.flush()
        await self.db.refresh(questionnaire)

        wb.close()
        return questionnaire, questions

    # =========================================================================
    # Analysis / Suggestion Generation
    # =========================================================================

    async def generate_suggestions(
        self,
        questionnaire_id: str,
        workspace_id: str,
    ) -> tuple[int, int, list[str], list[SkippedQuestion]]:
        """Generate ReminderSuggestion records from parsed questions.

        Includes cross-questionnaire deduplication: skips questions whose
        normalized text already has an accepted suggestion or existing reminder
        in the same workspace.

        Returns (suggestions_generated, skipped_count, domains_covered, skipped_details).
        """
        # Load questionnaire and questions
        stmt = (
            select(QuestionnaireResponse)
            .where(QuestionnaireResponse.id == questionnaire_id)
            .where(QuestionnaireResponse.workspace_id == workspace_id)
        )
        result = await self.db.execute(stmt)
        questionnaire = result.scalar_one_or_none()
        if not questionnaire:
            raise QuestionnaireServiceError(f"Questionnaire {questionnaire_id} not found")

        q_stmt = (
            select(QuestionnaireQuestion)
            .where(QuestionnaireQuestion.questionnaire_response_id == questionnaire_id)
            .order_by(QuestionnaireQuestion.source_row)
        )
        q_result = await self.db.execute(q_stmt)
        questions = list(q_result.scalars().all())

        # Delete previous pending suggestions for idempotent regeneration
        await self.db.execute(
            delete(ReminderSuggestion)
            .where(ReminderSuggestion.questionnaire_response_id == questionnaire_id)
            .where(ReminderSuggestion.status == "pending")
        )

        # Build dedup index: existing accepted/pending suggestions + reminders
        existing_keys = await self._build_dedup_index(workspace_id, questionnaire_id)

        suggestions_count = 0
        skipped_count = 0
        skipped_details: list[SkippedQuestion] = []
        domains_seen = set()
        # Track keys within this batch to avoid intra-questionnaire dupes
        batch_keys: set[str] = set()

        for q in questions:
            # Skip section headers
            if q.is_section_header:
                skipped_count += 1
                skipped_details.append(SkippedQuestion(
                    question_text=q.question_text, domain=q.domain, reason="header",
                ))
                continue

            # Skip blank responses
            if not q.response_text or q.response_text.strip() == "":
                skipped_count += 1
                skipped_details.append(SkippedQuestion(
                    question_text=q.question_text, domain=q.domain, reason="blank",
                ))
                continue

            response_lower = q.response_text.strip().lower()

            # Skip negative responses
            if response_lower in {"no", "never", "n/a", "na", "not applicable", "none"}:
                skipped_count += 1
                skipped_details.append(SkippedQuestion(
                    question_text=q.question_text, domain=q.domain, reason="negative",
                ))
                continue

            # Deduplication check
            dedup_key = self._normalize_question_key(q.question_text, q.domain)
            if dedup_key in existing_keys or dedup_key in batch_keys:
                logger.debug(f"Skipping duplicate question: {q.question_text[:60]}")
                skipped_count += 1
                entry = existing_keys.get(dedup_key)
                skipped_details.append(SkippedQuestion(
                    question_text=q.question_text,
                    domain=q.domain,
                    reason="duplicate",
                    duplicate_of_id=entry.item_id if entry and entry.item_id else None,
                    duplicate_of_type=entry.item_type if entry else None,
                    duplicate_of_title=entry.title if entry else None,
                ))
                continue

            # Generate suggestion based on response type
            suggestion = self._create_suggestion_from_question(q, workspace_id, questionnaire_id)
            if suggestion:
                self.db.add(suggestion)
                suggestions_count += 1
                batch_keys.add(dedup_key)
                if q.domain:
                    domains_seen.add(q.domain)
            else:
                skipped_count += 1
                skipped_details.append(SkippedQuestion(
                    question_text=q.question_text, domain=q.domain, reason="no_suggestion",
                ))

        # Update questionnaire status
        questionnaire.total_suggestions_generated = suggestions_count
        questionnaire.status = "analyzed"

        await self.db.flush()

        return suggestions_count, skipped_count, sorted(domains_seen), skipped_details

    # =========================================================================
    # Read Operations
    # =========================================================================

    async def list_questionnaires(
        self,
        workspace_id: str,
    ) -> list[QuestionnaireResponse]:
        """List all questionnaires for a workspace."""
        stmt = (
            select(QuestionnaireResponse)
            .where(QuestionnaireResponse.workspace_id == workspace_id)
            .order_by(QuestionnaireResponse.created_at.desc())
        )
        result = await self.db.execute(stmt)
        return list(result.scalars().all())

    async def get_questionnaire(
        self,
        questionnaire_id: str,
        workspace_id: str,
    ) -> QuestionnaireResponse | None:
        """Get a single questionnaire by ID."""
        stmt = (
            select(QuestionnaireResponse)
            .where(QuestionnaireResponse.id == questionnaire_id)
            .where(QuestionnaireResponse.workspace_id == workspace_id)
        )
        result = await self.db.execute(stmt)
        return result.scalar_one_or_none()

    async def get_questions(
        self,
        questionnaire_id: str,
    ) -> list[QuestionnaireQuestion]:
        """Get all questions for a questionnaire."""
        stmt = (
            select(QuestionnaireQuestion)
            .where(QuestionnaireQuestion.questionnaire_response_id == questionnaire_id)
            .order_by(QuestionnaireQuestion.source_row)
        )
        result = await self.db.execute(stmt)
        return list(result.scalars().all())

    async def delete_questionnaire(
        self,
        questionnaire_id: str,
        workspace_id: str,
    ) -> bool:
        """Delete a questionnaire (cascades to questions)."""
        stmt = (
            select(QuestionnaireResponse)
            .where(QuestionnaireResponse.id == questionnaire_id)
            .where(QuestionnaireResponse.workspace_id == workspace_id)
        )
        result = await self.db.execute(stmt)
        questionnaire = result.scalar_one_or_none()
        if not questionnaire:
            return False

        await self.db.delete(questionnaire)
        await self.db.flush()
        return True

    # =========================================================================
    # Private Helpers
    # =========================================================================

    @staticmethod
    def _normalize_question_key(question_text: str, domain: str | None = None) -> str:
        """Normalize a question into a dedup key.

        Strips punctuation, lowercases, collapses whitespace, and optionally
        prefixes with domain. Two questions with the same key are considered
        duplicates even if they come from different questionnaires.
        """
        text = question_text.lower().strip()
        text = re.sub(r'[^a-z0-9\s]', '', text)  # strip punctuation
        text = re.sub(r'\s+', ' ', text).strip()  # collapse whitespace
        prefix = (domain or "").lower().strip()
        return f"{prefix}::{text}" if prefix else text

    async def _build_dedup_index(self, workspace_id: str, exclude_questionnaire_id: str) -> dict[str, DedupEntry]:
        """Build a dict of normalized question keys → DedupEntry for existing items.

        Includes:
        - Accepted or pending suggestions from OTHER questionnaires
        - Questions from OTHER questionnaires that already generated suggestions
        - Existing reminders created from questionnaires

        Reminders take priority over suggestions (overwrite in dict).
        """
        from aexy.models.reminder import Reminder

        existing: dict[str, DedupEntry] = {}

        # 1. Existing suggestions (accepted or pending) from other questionnaires
        stmt = (
            select(
                ReminderSuggestion.id,
                ReminderSuggestion.suggested_title,
                ReminderSuggestion.suggested_domain,
            )
            .where(ReminderSuggestion.workspace_id == workspace_id)
            .where(ReminderSuggestion.questionnaire_response_id != exclude_questionnaire_id)
            .where(ReminderSuggestion.status.in_(["accepted", "pending"]))
        )
        result = await self.db.execute(stmt)
        for sid, title, domain in result.all():
            q_text = title
            if q_text.startswith("[") and "] " in q_text:
                q_text = q_text.split("] ", 1)[1]
            key = self._normalize_question_key(q_text, domain)
            existing[key] = DedupEntry(
                item_id=sid, item_type="suggestion", title=title, domain=domain,
            )

        # 2. Questions from other questionnaires (fallback dedup)
        q_stmt = (
            select(QuestionnaireQuestion.question_text, QuestionnaireQuestion.domain)
            .join(
                QuestionnaireResponse,
                QuestionnaireQuestion.questionnaire_response_id == QuestionnaireResponse.id,
            )
            .where(QuestionnaireResponse.workspace_id == workspace_id)
            .where(QuestionnaireResponse.id != exclude_questionnaire_id)
            .where(QuestionnaireQuestion.is_section_header == False)  # noqa: E712
        )
        q_result = await self.db.execute(q_stmt)
        for q_text, q_domain in q_result.all():
            key = self._normalize_question_key(q_text, q_domain)
            if key not in existing:
                existing[key] = DedupEntry(
                    item_id="", item_type="question", title=q_text, domain=q_domain,
                )

        # 3. Existing reminders (source_type = 'questionnaire') — highest priority
        r_stmt = (
            select(Reminder.id, Reminder.title, Reminder.domain)
            .where(Reminder.workspace_id == workspace_id)
            .where(Reminder.source_type == "questionnaire")
        )
        r_result = await self.db.execute(r_stmt)
        for rid, title, domain in r_result.all():
            q_text = title
            if q_text.startswith("[") and "] " in q_text:
                q_text = q_text.split("] ", 1)[1]
            key = self._normalize_question_key(q_text, domain)
            # Reminders overwrite suggestions (higher priority link)
            existing[key] = DedupEntry(
                item_id=rid, item_type="reminder", title=title, domain=domain,
            )

        logger.info(f"Dedup index built with {len(existing)} existing keys")
        return existing

    def _extract_metadata(self, wb) -> dict:
        """Extract metadata from the Document Summary sheet."""
        metadata = {}

        # Try the first sheet for key-value pairs
        if len(wb.sheetnames) == 0:
            return metadata

        sheet = wb[wb.sheetnames[0]]
        for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
            if row and len(row) >= 2 and row[0] and row[1]:
                key = str(row[0]).strip().lower()
                value = str(row[1]).strip()
                if not value or value == "None":
                    continue

                # Map common keys
                if any(k in key for k in ["partner", "vendor", "company", "organization"]):
                    metadata["partner_name"] = value
                elif any(k in key for k in ["year", "period", "date"]):
                    metadata["assessment_year"] = value
                elif any(k in key for k in ["title", "name", "document"]):
                    metadata["title"] = value
                elif any(k in key for k in ["version"]):
                    metadata["version"] = value
                else:
                    # Store any other key-value pairs
                    safe_key = re.sub(r'[^a-z0-9_]', '_', key)[:50]
                    metadata[safe_key] = value

        return metadata

    def _find_questions_sheet(self, wb):
        """Find the sheet containing questionnaire questions.

        Scans all sheets for rows matching header patterns. Prefers sheets
        with more header matches. Skips sheets that look like summaries
        (< 5 data rows).
        """
        best_sheet = None
        best_score = 0

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Check first 5 rows for header patterns
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                if row:
                    row_text = " ".join(str(v).lower() for v in row if v)
                    matches = sum(1 for pattern in HEADER_PATTERNS if pattern in row_text)
                    if matches > best_score:
                        best_score = matches
                        best_sheet = sheet

        # Need at least 2 pattern matches
        return best_sheet if best_score >= 2 else None

    def _detect_columns_exact(self, sheet) -> dict:
        """Tier 1: Detect columns via exact alias matching."""
        col_map = {}

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            if not row:
                continue

            found_count = 0
            temp_map = {}

            for col_idx, cell_value in enumerate(row):
                if cell_value is None:
                    continue
                cell_text = str(cell_value).strip().lower()

                for field_name, aliases in COLUMN_ALIASES.items():
                    if cell_text in aliases:
                        temp_map[field_name] = col_idx
                        found_count += 1
                        break

            # Accept if we found at least 2 columns (including question)
            if found_count >= 2 and "question" in temp_map:
                col_map = temp_map
                col_map["header_row"] = row_idx
                return col_map

        return col_map

    def _detect_columns_fuzzy(self, sheet) -> dict:
        """Tier 2: Detect columns via substring/fuzzy keyword matching."""
        col_map = {}

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            if not row:
                continue

            found_count = 0
            temp_map = {}
            used_cols = set()

            for field_name, keywords in FUZZY_KEYWORDS.items():
                for col_idx, cell_value in enumerate(row):
                    if cell_value is None or col_idx in used_cols:
                        continue
                    cell_text = str(cell_value).strip().lower()
                    if any(kw in cell_text for kw in keywords):
                        temp_map[field_name] = col_idx
                        used_cols.add(col_idx)
                        found_count += 1
                        break

            if found_count >= 2 and "question" in temp_map:
                col_map = temp_map
                col_map["header_row"] = row_idx
                logger.info(f"Column mapping resolved via fuzzy matching (row {row_idx}): {temp_map}")
                return col_map

        return col_map

    async def _detect_columns_llm(self, sheet, workspace_id: str | None = None) -> dict:
        """Tier 3: Use LLM to map columns when auto-detection fails."""
        gateway = self._get_llm_gateway()
        if not gateway:
            logger.warning("LLM gateway not available for column mapping fallback")
            return {}

        # Find the best candidate header row (row with most non-empty cells)
        best_row = None
        best_row_idx = 1
        best_count = 0

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            if not row:
                continue
            non_empty = sum(1 for v in row if v is not None and str(v).strip())
            if non_empty > best_count:
                best_count = non_empty
                best_row = row
                best_row_idx = row_idx

        if not best_row or best_count < 2:
            return {}

        # Build headers description for the LLM
        headers_desc = []
        for idx, val in enumerate(best_row):
            if val is not None and str(val).strip():
                headers_desc.append(f"  Column {idx}: \"{str(val).strip()}\"")

        headers_str = "\n".join(headers_desc)
        user_prompt = LLM_COLUMN_MAPPING_PROMPT.format(headers=headers_str)

        try:
            result = await gateway.call_llm(
                system_prompt="You are a data analyst. Respond only with valid JSON.",
                user_prompt=user_prompt,
                tokens_estimate=500,
                workspace_id=workspace_id,
            )
            response_text = result[0] if isinstance(result, tuple) else result

            # Parse JSON response — strip markdown fences if present
            clean = response_text.strip()
            if clean.startswith("```"):
                clean = clean.split("\n", 1)[-1].rsplit("```", 1)[0].strip()

            mapping = json.loads(clean)
            col_map = {}
            for field_name in ["serial_number", "domain", "question", "response", "possible_responses", "explanation"]:
                val = mapping.get(field_name)
                if val is not None and isinstance(val, int):
                    col_map[field_name] = val

            if "question" in col_map:
                col_map["header_row"] = best_row_idx
                logger.info(f"Column mapping resolved via LLM fallback (row {best_row_idx}): {col_map}")
                return col_map

        except Exception as e:
            logger.warning(f"LLM column mapping failed: {e}")

        return {}

    async def _detect_columns_with_fallback(self, sheet, workspace_id: str | None = None) -> dict:
        """3-tier column detection: exact alias → fuzzy substring → LLM.

        Returns a dict mapping field names to column indices, plus 'header_row'.
        """
        # Tier 1: Exact alias match
        col_map = self._detect_columns_exact(sheet)
        if col_map and "question" in col_map:
            logger.info(f"Column mapping resolved via exact alias match: {col_map}")
            return col_map

        # Tier 2: Fuzzy substring match
        col_map = self._detect_columns_fuzzy(sheet)
        if col_map and "question" in col_map:
            return col_map

        # Tier 3: LLM fallback
        col_map = await self._detect_columns_llm(sheet, workspace_id)
        if col_map and "question" in col_map:
            return col_map

        return col_map

    def _get_col_value(self, row_values: list, col_map: dict, field_name: str):
        """Get a column value from a row using the detected column map."""
        col_idx = col_map.get(field_name)
        if col_idx is None or col_idx >= len(row_values):
            return None
        return row_values[col_idx]

    def _is_section_header(
        self,
        question_text: str | None,
        response_text: str | None,
        possible_responses: str | None,
    ) -> bool:
        """Detect if a row is a section header rather than a question."""
        if not question_text:
            return False
        has_response = response_text and str(response_text).strip()
        has_options = possible_responses and str(possible_responses).strip()
        # Section headers typically have text but no response and no options
        return not has_response and not has_options

    def _classify_response_type(
        self,
        possible_responses: str | None,
        response_text: str | None,
    ) -> str:
        """Classify the response type based on possible responses."""
        if not possible_responses:
            return "text"

        options_lower = str(possible_responses).strip().lower()

        # Check for Yes/No pattern
        yes_no_patterns = {"yes\nno", "yes / no", "yes/no", "yes\r\nno", "yes, no"}
        if options_lower in yes_no_patterns or set(options_lower.split("\n")) == {"yes", "no"}:
            return "yes_no"

        # Check for frequency pattern
        if any(kw in options_lower for kw in FREQUENCY_KEYWORDS):
            return "frequency"

        # Multiple options = multi_choice
        separators = ["\n", "\r\n", " / ", "/"]
        for sep in separators:
            if sep in str(possible_responses) and len(str(possible_responses).split(sep)) > 2:
                return "multi_choice"

        return "text"

    def _create_suggestion_from_question(
        self,
        question: QuestionnaireQuestion,
        workspace_id: str,
        questionnaire_id: str,
    ) -> ReminderSuggestion | None:
        """Create a ReminderSuggestion from a question based on its response."""
        response_lower = question.response_text.strip().lower() if question.response_text else ""
        question_lower = question.question_text.lower() if question.question_text else ""
        domain = question.domain or "General"

        frequency = None
        confidence = 0.5
        category = "compliance"

        if question.response_type == "frequency":
            # Map frequency response to reminder frequency
            for keyword, freq_value in FREQUENCY_MAP.items():
                if keyword in response_lower:
                    frequency = freq_value
                    confidence = 0.9
                    break
            if not frequency:
                return None

        elif question.response_type == "yes_no":
            if response_lower != "yes":
                return None
            # "Yes" on review/audit/training type questions → yearly
            has_indicator = any(kw in question_lower for kw in YEARLY_INDICATOR_KEYWORDS)
            frequency = "yearly"
            confidence = 0.85 if has_indicator else 0.75

            # Refine category
            if any(kw in question_lower for kw in {"security", "vulnerability", "penetration", "firewall"}):
                category = "security"
            elif any(kw in question_lower for kw in {"audit", "soc", "iso"}):
                category = "audit"
            elif any(kw in question_lower for kw in {"training", "awareness"}):
                category = "training"
            elif any(kw in question_lower for kw in {"review", "assessment"}):
                category = "review"

        elif question.response_type == "multi_choice":
            # For multi_choice, use yearly as default with lower confidence
            frequency = "yearly"
            confidence = 0.65

        else:
            # Free text responses — lower confidence, yearly default
            frequency = "yearly"
            confidence = 0.5

        # Build title from domain + question context
        title = self._derive_title(domain, question.question_text)

        # Build description with context
        description = (
            f"Domain: {domain}\n"
            f"Question: {question.question_text}\n"
            f"Response: {question.response_text}"
        )
        if question.explanation:
            description += f"\nExplanation: {question.explanation}"

        return ReminderSuggestion(
            id=str(uuid4()),
            workspace_id=workspace_id,
            questionnaire_response_id=questionnaire_id,
            question_id=question.id,
            answer_text=question.response_text,
            suggested_title=title,
            suggested_description=description,
            suggested_category=category,
            suggested_frequency=frequency,
            suggested_domain=domain,
            confidence_score=confidence,
            status="pending",
        )

    def _derive_title(self, domain: str, question_text: str) -> str:
        """Derive a concise title from domain and question text."""
        # Truncate question to a reasonable title length
        q = question_text.strip()
        if len(q) > 80:
            # Try to cut at a word boundary
            q = q[:80].rsplit(" ", 1)[0] + "..."

        return f"[{domain}] {q}"
