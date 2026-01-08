"""Export service for generating reports in various formats."""

import csv
import io
import json
import os
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from uuid import uuid4

from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from aexy.models.analytics import ExportJob
from aexy.schemas.analytics import (
    ExportRequest,
    ExportJobResponse,
    ExportFormat,
    ExportType,
    ExportStatus,
    DateRange,
)

# Optional imports for PDF and Excel
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
    )
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Default export directory
DEFAULT_EXPORT_DIR = Path(tempfile.gettempdir()) / "aexy_exports"


class ExportService:
    """Service for generating exportable reports in various formats."""

    def __init__(self, export_dir: Path | None = None):
        self.export_dir = export_dir or DEFAULT_EXPORT_DIR
        self.export_dir.mkdir(parents=True, exist_ok=True)

    # -------------------------------------------------------------------------
    # Export Job Management
    # -------------------------------------------------------------------------

    async def create_export_job(
        self,
        request: ExportRequest,
        requester_id: str,
        db: AsyncSession,
    ) -> ExportJob:
        """Create a new export job."""
        # Validate format availability
        if request.format == ExportFormat.PDF and not REPORTLAB_AVAILABLE:
            raise ValueError("PDF export requires reportlab. Install with: pip install reportlab")
        if request.format == ExportFormat.XLSX and not OPENPYXL_AVAILABLE:
            raise ValueError("Excel export requires openpyxl. Install with: pip install openpyxl")

        job = ExportJob(
            id=str(uuid4()),
            requested_by=requester_id,
            export_type=request.export_type.value,
            format=request.format.value,
            config=request.config or {},
            status=ExportStatus.PENDING.value,
            expires_at=datetime.utcnow() + timedelta(hours=24),  # 24 hour expiry
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)
        return job

    async def get_export_job(
        self,
        job_id: str,
        db: AsyncSession,
        requester_id: str | None = None,
    ) -> ExportJob | None:
        """Get an export job by ID."""
        stmt = select(ExportJob).where(ExportJob.id == job_id)
        if requester_id:
            stmt = stmt.where(ExportJob.requested_by == requester_id)
        result = await db.execute(stmt)
        return result.scalar_one_or_none()

    async def list_export_jobs(
        self,
        db: AsyncSession,
        requester_id: str,
        status: ExportStatus | None = None,
        limit: int = 20,
    ) -> list[ExportJob]:
        """List export jobs for a user."""
        conditions = [ExportJob.requested_by == requester_id]
        if status:
            conditions.append(ExportJob.status == status.value)

        stmt = (
            select(ExportJob)
            .where(and_(*conditions))
            .order_by(ExportJob.created_at.desc())
            .limit(limit)
        )
        result = await db.execute(stmt)
        return list(result.scalars().all())

    async def update_job_status(
        self,
        job_id: str,
        db: AsyncSession,
        status: ExportStatus,
        file_path: str | None = None,
        file_size: int | None = None,
        error_message: str | None = None,
    ) -> ExportJob | None:
        """Update export job status."""
        job = await self.get_export_job(job_id, db)
        if not job:
            return None

        job.status = status.value
        if file_path:
            job.file_path = file_path
        if file_size:
            job.file_size_bytes = file_size
        if error_message:
            job.error_message = error_message
        if status == ExportStatus.COMPLETED:
            job.completed_at = datetime.utcnow()

        await db.commit()
        await db.refresh(job)
        return job

    async def cleanup_expired_exports(self, db: AsyncSession) -> int:
        """Delete expired export jobs and their files."""
        now = datetime.utcnow()
        stmt = select(ExportJob).where(ExportJob.expires_at < now)
        result = await db.execute(stmt)
        expired_jobs = result.scalars().all()

        count = 0
        for job in expired_jobs:
            # Delete file if exists
            if job.file_path:
                try:
                    path = Path(job.file_path)
                    if path.exists():
                        path.unlink()
                except Exception:
                    pass

            await db.delete(job)
            count += 1

        await db.commit()
        return count

    # -------------------------------------------------------------------------
    # Export Execution
    # -------------------------------------------------------------------------

    async def process_export(
        self,
        job_id: str,
        db: AsyncSession,
        data: dict,
    ) -> ExportJob | None:
        """Process an export job with the provided data."""
        job = await self.get_export_job(job_id, db)
        if not job:
            return None

        try:
            # Mark as processing
            await self.update_job_status(job_id, db, ExportStatus.PROCESSING)

            format_type = ExportFormat(job.format)
            file_path: str

            if format_type == ExportFormat.CSV:
                file_path = await self._export_csv(job, data)
            elif format_type == ExportFormat.JSON:
                file_path = await self._export_json(job, data)
            elif format_type == ExportFormat.PDF:
                file_path = await self._export_pdf(job, data)
            elif format_type == ExportFormat.XLSX:
                file_path = await self._export_xlsx(job, data)
            else:
                raise ValueError(f"Unsupported export format: {format_type}")

            # Get file size
            file_size = os.path.getsize(file_path)

            # Mark as completed
            return await self.update_job_status(
                job_id, db, ExportStatus.COMPLETED,
                file_path=file_path,
                file_size=file_size,
            )

        except Exception as e:
            await self.update_job_status(
                job_id, db, ExportStatus.FAILED,
                error_message=str(e),
            )
            raise

    def get_download_path(self, job: ExportJob) -> Path | None:
        """Get the download path for a completed export."""
        if not job.file_path:
            return None
        path = Path(job.file_path)
        return path if path.exists() else None

    # -------------------------------------------------------------------------
    # Format-Specific Export Implementations
    # -------------------------------------------------------------------------

    async def _export_csv(self, job: ExportJob, data: dict) -> str:
        """Export data to CSV format."""
        filename = f"{job.id}.csv"
        file_path = self.export_dir / filename

        with open(file_path, "w", newline="") as f:
            writer = csv.writer(f)

            # Handle different data structures
            if "rows" in data:
                # Standard tabular data
                if "headers" in data:
                    writer.writerow(data["headers"])
                for row in data["rows"]:
                    writer.writerow(row)

            elif "developers" in data:
                # Developer list export
                headers = ["ID", "Username", "Email", "GitHub URL", "Skills", "Created At"]
                writer.writerow(headers)
                for dev in data["developers"]:
                    skills = ", ".join(dev.get("top_skills", []))
                    writer.writerow([
                        dev.get("id", ""),
                        dev.get("github_username", ""),
                        dev.get("email", ""),
                        dev.get("github_url", ""),
                        skills,
                        dev.get("created_at", ""),
                    ])

            elif "widgets" in data:
                # Report widget data
                writer.writerow(["Widget", "Type", "Data"])
                for widget_id, widget_data in data["widgets"].items():
                    writer.writerow([
                        widget_data.get("title", widget_id),
                        widget_data.get("type", ""),
                        json.dumps(widget_data.get("data", {})),
                    ])

            elif "skills" in data and "developers" not in data:
                # Skill heatmap data
                if "developer_skills" in data:
                    headers = ["Developer"] + data["skills"]
                    writer.writerow(headers)
                    for dev_data in data["developer_skills"]:
                        row = [dev_data["developer_name"]]
                        skill_values = {s["skill"]: s["value"] for s in dev_data["skills"]}
                        for skill in data["skills"]:
                            row.append(skill_values.get(skill, 0))
                        writer.writerow(row)

            elif "trends" in data or "developer_trends" in data:
                # Productivity trends
                trends = data.get("developer_trends", data.get("trends", []))
                if trends:
                    # Flatten trends to CSV
                    headers = ["Developer", "Metric", "Period", "Value"]
                    writer.writerow(headers)
                    for trend in trends:
                        dev_id = trend.get("developer_id", "")
                        for i, val in enumerate(trend.get("commits", [])):
                            writer.writerow([dev_id, "commits", i, val])
                        for i, val in enumerate(trend.get("prs_merged", [])):
                            writer.writerow([dev_id, "prs_merged", i, val])
                        for i, val in enumerate(trend.get("reviews", [])):
                            writer.writerow([dev_id, "reviews", i, val])

            else:
                # Generic dict export - flatten to key-value pairs
                writer.writerow(["Key", "Value"])
                self._flatten_dict_to_csv(writer, data)

        return str(file_path)

    def _flatten_dict_to_csv(self, writer, data: dict, prefix: str = "") -> None:
        """Recursively flatten a dict to CSV rows."""
        for key, value in data.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if isinstance(value, dict):
                self._flatten_dict_to_csv(writer, value, full_key)
            elif isinstance(value, list):
                writer.writerow([full_key, json.dumps(value)])
            else:
                writer.writerow([full_key, value])

    async def _export_json(self, job: ExportJob, data: dict) -> str:
        """Export data to JSON format."""
        filename = f"{job.id}.json"
        file_path = self.export_dir / filename

        export_data = {
            "export_info": {
                "job_id": job.id,
                "export_type": job.export_type,
                "generated_at": datetime.utcnow().isoformat(),
            },
            "data": data,
        }

        with open(file_path, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

        return str(file_path)

    async def _export_pdf(self, job: ExportJob, data: dict) -> str:
        """Export data to PDF format."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")

        filename = f"{job.id}.pdf"
        file_path = self.export_dir / filename

        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=30,
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
        )
        body_style = styles["Normal"]

        elements = []

        # Title
        title = data.get("title", data.get("report_name", "Aexy Export"))
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            body_style,
        ))
        elements.append(Spacer(1, 20))

        # Handle different data types
        if "widgets" in data:
            # Report with widgets
            for widget_id, widget_data in data["widgets"].items():
                elements.append(Paragraph(
                    widget_data.get("title", widget_id),
                    heading_style,
                ))

                widget_content = widget_data.get("data", {})
                if "error" in widget_content:
                    elements.append(Paragraph(
                        f"Error: {widget_content['error']}",
                        body_style,
                    ))
                else:
                    # Render as table if possible
                    table_data = self._data_to_table(widget_content)
                    if table_data:
                        table = Table(table_data)
                        table.setStyle(TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]))
                        elements.append(table)
                    else:
                        elements.append(Paragraph(
                            json.dumps(widget_content, indent=2),
                            body_style,
                        ))

                elements.append(Spacer(1, 20))

        elif "developers" in data:
            # Developer profile(s)
            elements.append(Paragraph("Developer Profiles", heading_style))
            for dev in data["developers"]:
                elements.append(Paragraph(
                    f"<b>{dev.get('github_username', 'Unknown')}</b>",
                    body_style,
                ))
                if dev.get("email"):
                    elements.append(Paragraph(f"Email: {dev['email']}", body_style))
                if dev.get("top_skills"):
                    elements.append(Paragraph(
                        f"Top Skills: {', '.join(dev['top_skills'][:5])}",
                        body_style,
                    ))
                elements.append(Spacer(1, 10))

        elif "developer_skills" in data:
            # Skill heatmap
            elements.append(Paragraph("Skill Distribution", heading_style))
            table_data = [["Developer"] + data.get("skills", [])[:10]]
            for dev_data in data["developer_skills"]:
                row = [dev_data["developer_name"]]
                skill_values = {s["skill"]: s["value"] for s in dev_data["skills"]}
                for skill in data.get("skills", [])[:10]:
                    row.append(str(skill_values.get(skill, 0)))
                table_data.append(row)

            if len(table_data) > 1:
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(table)

        elif "risk_score" in data or "risk_level" in data:
            # Predictive insight
            elements.append(Paragraph("Risk Assessment", heading_style))
            if "risk_level" in data:
                elements.append(Paragraph(
                    f"Risk Level: <b>{data['risk_level'].upper()}</b>",
                    body_style,
                ))
            if "risk_score" in data:
                elements.append(Paragraph(
                    f"Risk Score: {data['risk_score']:.1%}",
                    body_style,
                ))
            if "factors" in data:
                elements.append(Paragraph("Contributing Factors:", body_style))
                for factor in data["factors"]:
                    elements.append(Paragraph(
                        f"  • {factor.get('factor', factor)}",
                        body_style,
                    ))
            if "recommendations" in data:
                elements.append(Paragraph("Recommendations:", body_style))
                for rec in data["recommendations"]:
                    elements.append(Paragraph(f"  • {rec}", body_style))

        else:
            # Generic data dump
            elements.append(Paragraph("Export Data", heading_style))
            table_data = self._data_to_table(data)
            if table_data:
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]))
                elements.append(table)
            else:
                elements.append(Paragraph(
                    json.dumps(data, indent=2, default=str),
                    body_style,
                ))

        doc.build(elements)
        return str(file_path)

    def _data_to_table(self, data: dict) -> list[list] | None:
        """Convert dict data to table rows if possible."""
        if not data:
            return None

        # Handle rows with headers
        if "rows" in data and "headers" in data:
            return [data["headers"]] + data["rows"]

        # Handle list of dicts
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            headers = list(data[0].keys())
            rows = [headers]
            for item in data[:50]:  # Limit rows
                rows.append([str(item.get(h, "")) for h in headers])
            return rows

        # Handle simple key-value dict
        if all(not isinstance(v, (dict, list)) for v in data.values()):
            return [["Key", "Value"]] + [[k, str(v)] for k, v in data.items()]

        return None

    async def _export_xlsx(self, job: ExportJob, data: dict) -> str:
        """Export data to Excel format."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        filename = f"{job.id}.xlsx"
        file_path = self.export_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row_num = 1

        # Title
        title = data.get("title", data.get("report_name", "Aexy Export"))
        ws.cell(row=row_num, column=1, value=title).font = Font(size=16, bold=True)
        row_num += 1
        ws.cell(row=row_num, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
        row_num += 2

        # Handle different data types
        if "widgets" in data:
            # Each widget as a section
            for widget_id, widget_data in data["widgets"].items():
                ws.cell(row=row_num, column=1, value=widget_data.get("title", widget_id)).font = Font(bold=True)
                row_num += 1

                widget_content = widget_data.get("data", {})
                if "error" not in widget_content:
                    row_num = self._write_data_to_sheet(ws, widget_content, row_num, header_font, header_fill, header_alignment, border)

                row_num += 1

        elif "developers" in data:
            # Developer table
            headers = ["Username", "Email", "GitHub URL", "Top Skills", "Created At"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            row_num += 1

            for dev in data["developers"]:
                ws.cell(row=row_num, column=1, value=dev.get("github_username", "")).border = border
                ws.cell(row=row_num, column=2, value=dev.get("email", "")).border = border
                ws.cell(row=row_num, column=3, value=dev.get("github_url", "")).border = border
                ws.cell(row=row_num, column=4, value=", ".join(dev.get("top_skills", [])[:5])).border = border
                ws.cell(row=row_num, column=5, value=dev.get("created_at", "")).border = border
                row_num += 1

        elif "developer_skills" in data:
            # Skill heatmap as table
            skills = data.get("skills", [])[:15]
            headers = ["Developer"] + skills
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            row_num += 1

            for dev_data in data["developer_skills"]:
                ws.cell(row=row_num, column=1, value=dev_data["developer_name"]).border = border
                skill_values = {s["skill"]: s["value"] for s in dev_data["skills"]}
                for col, skill in enumerate(skills, 2):
                    value = skill_values.get(skill, 0)
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border
                    # Color code based on value
                    if value >= 80:
                        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    elif value >= 60:
                        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                    elif value >= 40:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value > 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                row_num += 1

        elif "developer_trends" in data:
            # Productivity trends
            ws.title = "Productivity Trends"
            headers = ["Developer", "Total Commits", "Total PRs", "Total Reviews"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row_num += 1

            for trend in data["developer_trends"]:
                ws.cell(row=row_num, column=1, value=trend.get("developer_id", ""))
                ws.cell(row=row_num, column=2, value=sum(trend.get("commits", [])))
                ws.cell(row=row_num, column=3, value=sum(trend.get("prs_merged", [])))
                ws.cell(row=row_num, column=4, value=sum(trend.get("reviews", [])))
                row_num += 1

        else:
            # Generic data
            row_num = self._write_data_to_sheet(ws, data, row_num, header_font, header_fill, header_alignment, border)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        return str(file_path)

    def _write_data_to_sheet(self, ws, data: dict, row_num: int, header_font, header_fill, header_alignment, border) -> int:
        """Write dict data to worksheet and return next row number."""
        # Handle rows with headers
        if "rows" in data and "headers" in data:
            for col, header in enumerate(data["headers"], 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            row_num += 1

            for row in data["rows"]:
                for col, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col, value=value).border = border
                row_num += 1

        # Handle list of dicts
        elif isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row_num += 1

            for item in data[:100]:  # Limit rows
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col, value=str(item.get(header, "")))
                row_num += 1

        # Handle simple key-value dict
        else:
            ws.cell(row=row_num, column=1, value="Key").font = header_font
            ws.cell(row=row_num, column=2, value="Value").font = header_font
            row_num += 1

            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, default=str)
                ws.cell(row=row_num, column=1, value=key)
                ws.cell(row=row_num, column=2, value=str(value))
                row_num += 1

        return row_num

    # -------------------------------------------------------------------------
    # Convenience Methods
    # -------------------------------------------------------------------------

    async def export_developer_profile(
        self,
        developer_data: dict,
        format: ExportFormat,
        db: AsyncSession,
        requester_id: str,
    ) -> ExportJob:
        """Export a developer profile."""
        request = ExportRequest(
            export_type=ExportType.DEVELOPER_PROFILE,
            format=format,
            config={"developer_id": developer_data.get("id")},
        )
        job = await self.create_export_job(request, requester_id, db)
        return await self.process_export(job.id, db, {"developers": [developer_data]})

    async def export_team_analytics(
        self,
        analytics_data: dict,
        format: ExportFormat,
        db: AsyncSession,
        requester_id: str,
    ) -> ExportJob:
        """Export team analytics data."""
        request = ExportRequest(
            export_type=ExportType.TEAM_ANALYTICS,
            format=format,
            config={},
        )
        job = await self.create_export_job(request, requester_id, db)
        return await self.process_export(job.id, db, analytics_data)

    async def export_report(
        self,
        report_data: dict,
        format: ExportFormat,
        db: AsyncSession,
        requester_id: str,
    ) -> ExportJob:
        """Export a custom report."""
        request = ExportRequest(
            export_type=ExportType.REPORT,
            format=format,
            config={"report_id": report_data.get("report_id")},
        )
        job = await self.create_export_job(request, requester_id, db)
        return await self.process_export(job.id, db, report_data)


# Convenience function
def get_export_service(export_dir: Path | None = None) -> ExportService:
    """Get an instance of the export service."""
    return ExportService(export_dir=export_dir)
